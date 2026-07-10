"""将校对结果导出为 Excel 问题清单。"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from proofreader.checkers.consistency_checker import IssueLevel
from proofreader.pipeline import ProofreadBatchResult, ProofreadResult


def _level_text(level: IssueLevel) -> str:
    return {
        IssueLevel.ERROR: "严重",
        IssueLevel.WARNING: "警告",
        IssueLevel.INFO: "提示",
    }.get(level, "未知")


def export_to_excel(result: ProofreadResult, output_path: Path | str) -> Path:
    """将各类问题汇总导出为 Excel 工作簿。"""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # 一致性/偏离问题
    ws1 = wb.active
    ws1.title = "一致性偏离"
    ws1.append(["问题ID", "级别", "类型", "需求条目", "投标响应", "问题描述", "建议"])
    for issue in result.consistency_issues:
        ws1.append([
            issue.issue_id,
            _level_text(issue.level),
            issue.issue_type.value,
            issue.requirement_text,
            issue.bid_text,
            issue.message,
            issue.suggestion,
        ])

    # 表格问题
    ws2 = wb.create_sheet("表格问题")
    ws2.append(["问题ID", "需求表索引", "投标表索引", "问题描述", "建议", "差异详情"])
    for table_issue in result.table_issues:
        ws2.append([
            table_issue.issue_id,
            table_issue.requirement_table_index,
            table_issue.bid_table_index,
            table_issue.message,
            table_issue.suggestion,
            " | ".join(table_issue.details[:20]),
        ])

    # 错别字
    ws3 = wb.create_sheet("错别字")
    ws3.append(["错词", "建议改为", "位置", "消息"])
    for typo in result.typo_issues:
        ws3.append([
            typo.word,
            typo.suggestion,
            typo.position,
            typo.message,
        ])

    # OCR 问题
    ws4 = wb.create_sheet("截图OCR")
    ws4.append(["图片编号", "上下文", "缺失关键词", "消息"])
    for ocr in result.ocr_issues:
        ws4.append([
            ocr.image_index,
            ocr.context_block.text[:100] if ocr.context_block else "",
            ", ".join(ocr.missing_keywords[:10]),
            ocr.message,
        ])

    # 统一表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="343A40", end_color="343A40", fill_type="solid")
    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # 自动调整列宽（粗略）
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except Exception:
                    pass
            adjusted_width = min(max(max_length + 2, 12), 60)
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)
    return output_path


def _safe_sheet_name(bid_name: str) -> str:
    """生成合法的 Excel sheet 名称（长度 ≤ 31，不含非法字符）。"""
    # Excel sheet 名不能包含 : \ / ? * [ ]
    safe = "".join(c for c in bid_name if c not in r':\/?*[]')
    return safe[:31]


def export_batch_to_excel(batch_result: ProofreadBatchResult, output_path: Path | str) -> Path:
    """将批量校对结果导出为 Excel，每份投标文件一个工作表。"""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # 删除默认 sheet，按实际投标文件创建
    wb.remove(wb.active)

    if not batch_result.items:
        wb.create_sheet("无结果")
        wb.save(output_path)
        return output_path

    for item in batch_result.items:
        sheet_name = _safe_sheet_name(item.bid_path.stem)
        # 处理重名 sheet
        base_name = sheet_name
        counter = 1
        while sheet_name in wb.sheetnames:
            suffix = f"_{counter}"
            sheet_name = (base_name[: 31 - len(suffix)] + suffix)
            counter += 1

        ws = wb.create_sheet(sheet_name)
        req_names = "、".join(p.name for p in item.requirement_paths)
        ws.append(["需求文件", "投标文件"])
        ws.append([req_names, item.bid_path.name])
        ws.append([])

        # 合并四类问题到同一个 sheet，增加「类型」列
        ws.append(["问题ID", "类型", "级别", "需求条目/上下文", "投标响应/错词", "问题描述", "建议"])
        for issue in item.result.consistency_issues:
            ws.append([
                issue.issue_id,
                "一致性偏离",
                _level_text(issue.level),
                issue.requirement_text,
                issue.bid_text,
                issue.message,
                issue.suggestion,
            ])
        for table_issue in item.result.table_issues:
            ws.append([
                table_issue.issue_id,
                "表格问题",
                "",
                table_issue.message,
                "",
                " | ".join(table_issue.details[:20]),
                table_issue.suggestion,
            ])
        for typo in item.result.typo_issues:
            ws.append([
                "",
                "错别字",
                "",
                typo.word,
                typo.suggestion,
                typo.message,
                "",
            ])
        for ocr in item.result.ocr_issues:
            ws.append([
                f"图片 #{ocr.image_index}",
                "截图OCR",
                "",
                ocr.context_block.text[:100] if ocr.context_block else "",
                "",
                ocr.message,
                "",
            ])

    # 表头样式：批量 sheet 的真正的表头位于第 4 行（前 3 行为元数据空行）
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="343A40", end_color="343A40", fill_type="solid")
    for ws in wb.worksheets:
        header_row = 4
        for cell in ws[header_row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except Exception:
                    pass
            adjusted_width = min(max(max_length + 2, 12), 60)
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)
    return output_path
