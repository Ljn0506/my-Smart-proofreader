"""pytest 风格的报告导出测试。"""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from proofreader.exporters import export_batch_to_excel, export_to_excel
from proofreader.pipeline import ProofreadBatchResult, ProofreadResult


def test_export_to_excel_creates_file(sample_result: ProofreadResult) -> None:
    """Excel 导出应生成可读取的 xlsx 文件。"""
    output = Path("/tmp/test_report.xlsx")
    export_to_excel(sample_result, output)

    assert output.exists()
    wb = load_workbook(output)
    sheet_names = wb.sheetnames
    assert "一致性偏离" in sheet_names
    assert "表格问题" in sheet_names
    assert "错别字" in sheet_names
    assert "截图OCR" in sheet_names

    # 检查一致性偏离 sheet 有数据
    ws = wb["一致性偏离"]
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) >= 2  # 表头 + 至少一行数据

    output.unlink(missing_ok=True)


def test_export_batch_to_excel_creates_file(sample_result: ProofreadResult, tmp_path: Path) -> None:
    """批量 Excel 导出应为每份投标文件生成一个工作表。"""
    from proofreader.pipeline import BatchResultItem, ProofreadBatchResult

    batch = ProofreadBatchResult(
        items=[
            BatchResultItem(Path("bid.docx"), sample_result, [Path("requirements.docx")]),
            BatchResultItem(Path("bid_copy.docx"), sample_result, [Path("requirements_copy.docx")]),
        ],
        errors=[],
    )
    output = tmp_path / "batch_report.xlsx"
    export_batch_to_excel(batch, output)
    assert output.exists()

    wb = load_workbook(output)
    assert len(wb.sheetnames) == 2
